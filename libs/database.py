"""
database.py v14 — Source de vérité UNIQUE (tout en SQL, zéro JSON métier)

Modèle conceptuel v9 :
  Robot   = instance nommée (robot1, robot2...) liée à 1 compte Facebook
  Media   = fichier image/vidéo avec captcha optionnel + description propre
  Comment = texte stocké en DB, tiré aléatoirement par robot
  DM      = file d'attente de messages privés (amis / abonnés page)

Nouvelles tables v9 :
  robots              — instances nommées (robot1..N), 1 robot = 1 compte FB
  robot_groups        — groupes assignés à un robot
  robot_media         — médias assignés à un robot (many-to-many)
  robot_campaigns     — campagnes assignées à un robot
  media_assets        — médias avec captcha optionnel + description propre
  comments            — commentaires stockés en DB, tirés aléatoirement
  publications        — historique complet (anti-doublon intégré)
  published_comments  — historique des commentaires publiés
  dm_queue            — file DM planifiés/en attente
  subscriptions       — abonnements groupe par compte
  circuit_breaker_state — état CB persistant par robot

CORRECTIONS v9 :
  - was_published_recently() : anti-doublon strict par (robot, groupe, variant)
  - pick_random_media()      : gère captcha → concatène description + captcha_text
  - get_groups_for_robot()   : groupes filtrés par robot (vs session générique)
  - add_robot()              : création/upsert d'un robot avec compte associé
"""

import sqlite3, json, threading, pathlib, csv
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Union

try:
    from libs.log_emitter import emit
except ImportError:
    from log_emitter import emit


class BONDatabase:

    def __init__(self, db_path=None):
        if db_path is None:
            try:
                from libs.config_manager import LOGS_DIR
                db_path = str(LOGS_DIR / "bon.db")
            except ImportError:
                db_path = "logs/bon.db"
        self.db_path = pathlib.Path(db_path)
        # Pour :memory:, mkdir échouerait sur le parent '.' → on skipppe
        if str(self.db_path) != ":memory:":
            self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.RLock()
        # Connexion persistante : indispensable pour :memory: (chaque
        # sqlite3.connect(':memory:') crée une base distincte et vide).
        # Pour les fichiers sur disque, une connexion longue durée est aussi
        # plus efficace (WAL activé une seule fois, cache partagé).
        self._conn = sqlite3.connect(
            str(self.db_path), check_same_thread=False
        )
        self._conn.row_factory = sqlite3.Row
        self._conn.execute("PRAGMA journal_mode=WAL")
        self._conn.execute("PRAGMA foreign_keys=ON")
        self._conn.execute("PRAGMA synchronous=NORMAL")
        self._conn.execute("PRAGMA cache_size=-8000")
        self._closed = False
        self._init_db()

    # ── Cycle de vie ───────────────────────────────────────────────────────

    def close(self) -> None:
        """Ferme explicitement la connexion SQLite.

        A appeler dans les tests apres chaque instance BONDatabase(:memory:)
        ou en fin de script pour liberer le descripteur de fichier.
        Idempotent : appels multiples sans effet.
        """
        with self._lock:
            if not self._closed:
                try:
                    self._conn.close()
                finally:
                    self._closed = True

    def __del__(self) -> None:
        """Filet de securite : ferme la connexion si close() n'a pas ete appele."""
        try:
            self.close()
        except Exception:
            pass

    # ── Connexion ──────────────────────────────────────────────────────────

    def _connect(self):
        """Retourne la connexion persistante (thread-safe via self._lock)."""
        if self._closed:
            raise RuntimeError("BONDatabase: connexion deja fermee (close() appele)")
        return self._conn

    def _exec_conn(self, conn, sql, params=()):
        return conn.execute(sql, params)

    def _query_conn(self, conn, sql, params=()):
        return [dict(r) for r in conn.execute(sql, params).fetchall()]

    def _scalar_conn(self, conn, sql, params=()):
        r = conn.execute(sql, params).fetchone()
        return r[0] if r else None

    def _exec(self, sql, params=()):
        with self._lock:
            conn = self._connect()
            cur = conn.execute(sql, params)
            conn.commit()
            return cur

    def _query(self, sql, params=()):
        with self._lock:
            conn = self._connect()
            return self._query_conn(conn, sql, params)

    def _query_one(self, sql, params=()):
        r = self._query(sql, params)
        return r[0] if r else None

    def _query_scalar(self, sql, params=()):
        with self._lock:
            conn = self._connect()
            return self._scalar_conn(conn, sql, params)

    # ── Schéma complet v9 ─────────────────────────────────────────────────

    def _init_db(self):
        ddl = [
            # Comptes Facebook (1 compte = 1 identité FB)
            """CREATE TABLE IF NOT EXISTS accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                email TEXT, profile_url TEXT,
                health_score INTEGER DEFAULT 100,
                status TEXT DEFAULT 'healthy',
                status_reason TEXT,
                total_posts INTEGER DEFAULT 0,
                successful_posts INTEGER DEFAULT 0,
                failed_posts INTEGER DEFAULT 0,
                blocked_count INTEGER DEFAULT 0,
                consecutive_failures INTEGER DEFAULT 0,
                last_post_date TEXT, last_activity_date TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                max_groups_per_day INTEGER DEFAULT 20,
                cooldown_until TEXT,
                warmup_completed INTEGER DEFAULT 0)""",

            # Robots (instances nommées : robot1, robot2...)
            # 1 robot = 1 compte FB + 1 storage_state Playwright
            """CREATE TABLE IF NOT EXISTS robots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                robot_name TEXT UNIQUE NOT NULL,
                account_id INTEGER NOT NULL,
                storage_state_path TEXT NOT NULL,
                max_groups_per_run INTEGER DEFAULT 10,
                max_groups_per_hour INTEGER DEFAULT 5,
                delay_min_s INTEGER DEFAULT 60,
                delay_max_s INTEGER DEFAULT 120,
                max_runs_per_day INTEGER DEFAULT 2,
                cooldown_between_runs_s INTEGER DEFAULT 7200,
                locale TEXT DEFAULT 'fr-FR',
                timezone_id TEXT DEFAULT 'Europe/Paris',
                platform TEXT DEFAULT 'windows',
                proxy_server TEXT,
                proxy_username TEXT,
                proxy_password TEXT,
                telegram_token TEXT DEFAULT '',
                telegram_chat_id TEXT DEFAULT '',
                captcha_api_key TEXT DEFAULT NULL,
                active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (account_id) REFERENCES accounts(id))""",

            # Stats de runs par robot (anti-doublon run journalier)
            """CREATE TABLE IF NOT EXISTS robot_run_stats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                robot_name TEXT NOT NULL,
                run_date TEXT NOT NULL,
                run_count INTEGER DEFAULT 0,
                last_run_ts TEXT,
                UNIQUE(robot_name, run_date))""",

            # Groupes FB
            """CREATE TABLE IF NOT EXISTS groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT UNIQUE NOT NULL,
                name TEXT, category TEXT,
                language TEXT DEFAULT 'fr',
                quality_score INTEGER DEFAULT 50,
                members_count INTEGER,
                activity_level TEXT,
                total_posts INTEGER DEFAULT 0,
                successful_posts INTEGER DEFAULT 0,
                failed_posts INTEGER DEFAULT 0,
                rejected_posts INTEGER DEFAULT 0,
                last_post_date TEXT,
                first_seen_date TEXT DEFAULT CURRENT_TIMESTAMP,
                active INTEGER DEFAULT 1)""",

            # Groupes assignés à un robot (many-to-many)
            """CREATE TABLE IF NOT EXISTS robot_groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                robot_id INTEGER NOT NULL,
                group_id INTEGER NOT NULL,
                active INTEGER DEFAULT 1,
                UNIQUE(robot_id, group_id),
                FOREIGN KEY (robot_id) REFERENCES robots(id),
                FOREIGN KEY (group_id) REFERENCES groups(id))""",

            # Campagnes
            """CREATE TABLE IF NOT EXISTS campaigns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                description TEXT,
                active INTEGER DEFAULT 1,
                language TEXT DEFAULT 'fr',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",

            # Variantes de campagne (text seulement — médias liés séparément)
            """CREATE TABLE IF NOT EXISTS campaign_variants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                campaign_id INTEGER NOT NULL,
                variant_key TEXT NOT NULL,
                text_fr TEXT, text_en TEXT, text_ar TEXT,
                cta TEXT,
                weight INTEGER DEFAULT 1,
                bg_color TEXT,
                post_type TEXT DEFAULT 'text_image',
                UNIQUE(campaign_id, variant_key),
                FOREIGN KEY (campaign_id) REFERENCES campaigns(id))""",

            # Campagnes assignées à un robot
            """CREATE TABLE IF NOT EXISTS robot_campaigns (
                robot_id INTEGER NOT NULL,
                campaign_id INTEGER NOT NULL,
                PRIMARY KEY (robot_id, campaign_id),
                FOREIGN KEY (robot_id) REFERENCES robots(id),
                FOREIGN KEY (campaign_id) REFERENCES campaigns(id))""",

            # Médias (images / vidéos)
            # captcha_text : texte optionnel lié à l'image (ex: watermark, code promo)
            #   → sera concaténé au hasard avec une description avant publication
            # description  : description libre de l'image (tirée aléatoirement si plusieurs)
            """CREATE TABLE IF NOT EXISTS media_assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_path TEXT NOT NULL,
                file_name TEXT NOT NULL,
                campaign_id INTEGER,
                captcha_text TEXT,
                description TEXT,
                active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (campaign_id) REFERENCES campaigns(id))""",

            # Médias assignés à un robot (many-to-many)
            """CREATE TABLE IF NOT EXISTS robot_media (
                robot_id INTEGER NOT NULL,
                media_id INTEGER NOT NULL,
                PRIMARY KEY (robot_id, media_id),
                FOREIGN KEY (robot_id) REFERENCES robots(id),
                FOREIGN KEY (media_id) REFERENCES media_assets(id))""",

            # Commentaires (bank de commentaires, tirés aléatoirement)
            """CREATE TABLE IF NOT EXISTS comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                text TEXT NOT NULL,
                robot_name TEXT,
                active INTEGER DEFAULT 1,
                used_count INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",

            # Publications (historique complet + anti-doublon)
            """CREATE TABLE IF NOT EXISTS publications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                robot_name TEXT NOT NULL,
                account_id INTEGER NOT NULL,
                group_id INTEGER NOT NULL,
                campaign_name TEXT,
                variant_id TEXT,
                text_content TEXT,
                images TEXT,
                bg_color TEXT,
                post_type TEXT DEFAULT 'text_image',
                status TEXT NOT NULL,
                error_message TEXT,
                screenshot_path TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                published_at TEXT,
                FOREIGN KEY (account_id) REFERENCES accounts(id),
                FOREIGN KEY (group_id) REFERENCES groups(id))""",

            # Commentaires publiés (historique)
            """CREATE TABLE IF NOT EXISTS published_comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                robot_name TEXT NOT NULL,
                account_id INTEGER NOT NULL,
                group_id INTEGER,
                publication_id INTEGER,
                comment_text TEXT NOT NULL,
                status TEXT DEFAULT 'success',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (account_id) REFERENCES accounts(id))""",

            # Erreurs
            """CREATE TABLE IF NOT EXISTS errors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                robot_name TEXT,
                account_id INTEGER,
                group_id INTEGER,
                error_type TEXT NOT NULL,
                error_message TEXT,
                step TEXT,
                selector_key TEXT,
                screenshot_path TEXT,
                html_snapshot_path TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (account_id) REFERENCES accounts(id))""",

            # Statistiques sélecteurs
            """CREATE TABLE IF NOT EXISTS selector_stats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                selector_key TEXT UNIQUE NOT NULL,
                working_selector TEXT,
                total_attempts INTEGER DEFAULT 0,
                successful_attempts INTEGER DEFAULT 0,
                failed_attempts INTEGER DEFAULT 0,
                success_rate REAL DEFAULT 100.0,
                last_success_date TEXT,
                last_failure_date TEXT,
                last_failure_reason TEXT,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP)""",

            # Blocages comptes
            """CREATE TABLE IF NOT EXISTS account_blocks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_id INTEGER NOT NULL,
                block_type TEXT NOT NULL,
                reason TEXT,
                duration_hours INTEGER,
                started_at TEXT DEFAULT CURRENT_TIMESTAMP,
                ended_at TEXT,
                resolved INTEGER DEFAULT 0,
                FOREIGN KEY (account_id) REFERENCES accounts(id))""",

            # Abonnements groupe par compte
            """CREATE TABLE IF NOT EXISTS subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_id INTEGER NOT NULL,
                group_id INTEGER NOT NULL,
                status TEXT DEFAULT 'subscribed',
                subscribed_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(account_id, group_id),
                FOREIGN KEY (account_id) REFERENCES accounts(id),
                FOREIGN KEY (group_id) REFERENCES groups(id))""",

            # Circuit breaker persistant par robot
            """CREATE TABLE IF NOT EXISTS circuit_breaker_state (
                robot_name TEXT PRIMARY KEY,
                state TEXT DEFAULT 'closed',
                failures INTEGER DEFAULT 0,
                successes INTEGER DEFAULT 0,
                opened_at TEXT,
                recovery_timeout_s INTEGER DEFAULT 900,
                failure_threshold INTEGER DEFAULT 3,
                half_open_max_ok INTEGER DEFAULT 2,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP)""",

            # File DM planifiés
            """CREATE TABLE IF NOT EXISTS dm_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                robot_name TEXT NOT NULL,
                target_type TEXT NOT NULL,
                target_id TEXT NOT NULL,
                text_content TEXT NOT NULL,
                media_paths TEXT,
                status TEXT DEFAULT 'pending',
                attempts INTEGER DEFAULT 0,
                scheduled_at TEXT,
                sent_at TEXT,
                error_msg TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",

            # Config clé-valeur globale (Telegram global, CDN, etc.)
            """CREATE TABLE IF NOT EXISTS config_kv (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP)""",

            # Journal tentatives résolution CAPTCHA (v14)
            """CREATE TABLE IF NOT EXISTS captcha_solve_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                robot_name TEXT,
                solve_type TEXT,
                status TEXT,
                error_message TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",

            # Jobs planificateur APScheduler (v14)
            """CREATE TABLE IF NOT EXISTS scheduler_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT UNIQUE NOT NULL,
                robot_name TEXT NOT NULL,
                cron_expression TEXT NOT NULL,
                command_name TEXT DEFAULT 'post',
                active INTEGER DEFAULT 1,
                last_run_at TEXT,
                next_run_at TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",

            # Index
            "CREATE INDEX IF NOT EXISTS idx_pub_robot   ON publications(robot_name)",
            "CREATE INDEX IF NOT EXISTS idx_pub_account ON publications(account_id)",
            "CREATE INDEX IF NOT EXISTS idx_pub_group   ON publications(group_id)",
            "CREATE INDEX IF NOT EXISTS idx_pub_status  ON publications(status)",
            "CREATE INDEX IF NOT EXISTS idx_pub_date    ON publications(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_err_robot   ON errors(robot_name)",
            "CREATE INDEX IF NOT EXISTS idx_err_date    ON errors(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_run_robot   ON robot_run_stats(robot_name)",
            "CREATE INDEX IF NOT EXISTS idx_media_camp  ON media_assets(campaign_id)",
            "CREATE INDEX IF NOT EXISTS idx_sub_acc_grp ON subscriptions(account_id, group_id)",
            "CREATE INDEX IF NOT EXISTS idx_captcha_log_date ON captcha_solve_log(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_sched_robot ON scheduler_jobs(robot_name)",
        ]
        with self._lock:
            conn = self._connect()
            for stmt in ddl:
                conn.execute(stmt)
            conn.commit()
            self._apply_migrations(conn)
            self.bootstrap_default_workspace()
            emit("INFO", "DATABASE_INITIALIZED_V14", path=str(self.db_path))

    def _apply_migrations(self, conn):
        """Migrations idempotentes — silencieuses si colonne déjà présente."""
        migrations = [
            "ALTER TABLE accounts ADD COLUMN consecutive_failures INTEGER DEFAULT 0",
            "ALTER TABLE accounts ADD COLUMN warmup_completed INTEGER DEFAULT 0",
            "ALTER TABLE groups ADD COLUMN active INTEGER DEFAULT 1",
            "ALTER TABLE robots ADD COLUMN captcha_api_key TEXT DEFAULT NULL",
            "ALTER TABLE publications ADD COLUMN robot_name TEXT",
            "ALTER TABLE publications ADD COLUMN post_type TEXT DEFAULT 'text_image'",
            # v13 — Profil UI par compte (langue + variante interface Facebook)
            "ALTER TABLE accounts ADD COLUMN ui_lang TEXT DEFAULT NULL",
            "ALTER TABLE accounts ADD COLUMN ui_variant TEXT DEFAULT NULL",
            "ALTER TABLE accounts ADD COLUMN ui_confidence INTEGER DEFAULT 0",
            "ALTER TABLE accounts ADD COLUMN ui_detected_at TEXT DEFAULT NULL",
        ]
        for m in migrations:
            try:
                conn.execute(m)
                conn.commit()
            except Exception:
                pass

    def update_account_ui_profile(self, account, lang: str,
                                   variant: str, confidence: int = 0) -> None:
        """
        Sauvegarde le profil UI détecté pour un compte.
        Appelé par AccountUIProfiler après détection automatique.
        """
        account_id = self._resolve_account_id(account)
        if not account_id and isinstance(account, str):
            account_id = self.ensure_account_exists(account)
        if account_id:
            from datetime import datetime as _dt
            now = _dt.now().isoformat()
            self._exec(
                """UPDATE accounts
                   SET ui_lang=?, ui_variant=?, ui_confidence=?,
                       ui_detected_at=?, updated_at=?
                   WHERE id=?""",
                (lang, variant, confidence, now, now, account_id)
            )

    def get_ui_profile(self, account) -> dict:
        """
        Retourne le profil UI stocké pour un compte.
        Returns dict avec: ui_lang, ui_variant, ui_confidence, ui_detected_at

        Utilise dict(row) pour être safe sur des DB existantes où la migration
        _apply_migrations() n'a pas encore ajouté les colonnes ui_*.
        """
        row = (self.get_account(account) if isinstance(account, str)
               else self.get_account_by_id(account))
        if not row:
            return {}
        row_dict = dict(row) if row else {}
        result = {}
        for key in ("ui_lang", "ui_variant", "ui_confidence", "ui_detected_at"):
            result[key] = row_dict.get(key)
        return result

    # ── Helpers résolution ────────────────────────────────────────────────

    def _resolve_account_id(self, account):
        if isinstance(account, int):
            return account
        row = self._query_one("SELECT id FROM accounts WHERE name=?", (account,))
        return row["id"] if row else None

    def _resolve_group_id_conn(self, conn, group):
        if isinstance(group, int):
            return group
        rows = self._query_conn(conn, "SELECT id FROM groups WHERE url=?", (group,))
        if rows:
            return rows[0]["id"]
        cur = conn.execute("INSERT INTO groups (url) VALUES (?)", (group,))
        return cur.lastrowid

    def _resolve_robot_id(self, robot_name):
        row = self._query_one("SELECT id FROM robots WHERE robot_name=?", (robot_name,))
        return row["id"] if row else None

    # ══════════════════════════════════════════
    # ROBOTS
    # ══════════════════════════════════════════

    def upsert_robot(self, robot_name: str, account_name: str,
                     storage_state_path: str, config: dict = None) -> int:
        """
        Crée ou met à jour un robot.
        robot_name : identifiant unique (ex: 'robot1', 'robot2')
        account_name : nom du compte Facebook associé
        """
        config = config or {}
        account_id = self.ensure_account_exists(account_name)
        now = datetime.now().isoformat()

        proxy_server   = config.get("proxy_server")
        proxy_username = config.get("proxy_username")
        proxy_password = config.get("proxy_password")
        if isinstance(config.get("proxy"), dict):
            px = config["proxy"]
            proxy_server   = px.get("server", proxy_server)
            proxy_username = px.get("username", proxy_username)
            proxy_password = px.get("password", proxy_password)

        with self._lock:
            conn = self._connect()
            existing = self._scalar_conn(
                conn, "SELECT id FROM robots WHERE robot_name=?", (robot_name,)
            )
            delay = config.get("delay_between_groups", [60, 120])
            params = (
                storage_state_path,
                config.get("max_groups_per_run", 10),
                config.get("max_groups_per_hour", 5),
                delay[0] if isinstance(delay, list) else delay,
                delay[-1] if isinstance(delay, list) else delay,
                config.get("max_runs_per_day", 2),
                config.get("cooldown_between_runs_s", 7200),
                config.get("locale", "fr-FR"),
                config.get("timezone_id", "Europe/Paris"),
                config.get("platform", "windows"),
                proxy_server, proxy_username, proxy_password,
                config.get("telegram_token", ""),
                config.get("telegram_chat_id", ""),
                config.get("captcha_api_key", None),
                now,
            )
            if existing:
                conn.execute(
                    """UPDATE robots SET
                        account_id=?, storage_state_path=?,
                        max_groups_per_run=?, max_groups_per_hour=?,
                        delay_min_s=?, delay_max_s=?,
                        max_runs_per_day=?, cooldown_between_runs_s=?,
                        locale=?, timezone_id=?, platform=?,
                        proxy_server=?, proxy_username=?, proxy_password=?,
                        telegram_token=?, telegram_chat_id=?,
                        captcha_api_key=?, updated_at=?
                       WHERE robot_name=?""",
                    (account_id,) + params + (robot_name,)
                )
                conn.commit()
                return existing
            cur = conn.execute(
                """INSERT INTO robots (
                    robot_name, account_id, storage_state_path,
                    max_groups_per_run, max_groups_per_hour,
                    delay_min_s, delay_max_s,
                    max_runs_per_day, cooldown_between_runs_s,
                    locale, timezone_id, platform,
                    proxy_server, proxy_username, proxy_password,
                    telegram_token, telegram_chat_id, captcha_api_key,
                    created_at, updated_at
                   ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (robot_name, account_id) + params + (now,)
            )
            conn.commit()
            return cur.lastrowid

    def get_robot(self, robot_name: str) -> Optional[Dict]:
        row = self._query_one(
            """SELECT r.*, a.name AS account_name, a.health_score, a.status
               FROM robots r
               JOIN accounts a ON a.id = r.account_id
               WHERE r.robot_name=? AND r.active=1""",
            (robot_name,)
        )
        if not row:
            return None
        d = dict(row)
        # Reconstruire proxy dict pour compatibilité
        if d.get("proxy_server"):
            d["proxy"] = {
                "server":   d["proxy_server"],
                "username": d.get("proxy_username", ""),
                "password": d.get("proxy_password", ""),
            }
        else:
            d["proxy"] = None
        d["delay_between_groups"] = [d.get("delay_min_s", 60), d.get("delay_max_s", 120)]
        return d

    def get_all_robots(self) -> List[Dict]:
        return self._query(
            """SELECT r.*, a.name AS account_name, a.health_score, a.status
               FROM robots r JOIN accounts a ON a.id=r.account_id
               WHERE r.active=1 ORDER BY r.robot_name"""
        )

    def list_robot_names(self) -> List[str]:
        return [r["robot_name"] for r in self._query(
            "SELECT robot_name FROM robots WHERE active=1 ORDER BY robot_name"
        )]

    def list_sessions(self) -> List[str]:
        """Alias de list_robot_names() — rétrocompatibilité session_manager / config_manager."""
        return self.list_robot_names()

    def robot_exists(self, robot_name: str) -> bool:
        return bool(self._query_scalar(
            "SELECT COUNT(*) FROM robots WHERE robot_name=? AND active=1",
            (robot_name,)
        ))

    def delete_robot(self, robot_name: str) -> bool:
        try:
            self._exec("UPDATE robots SET active=0 WHERE robot_name=?", (robot_name,))
            return True
        except Exception as e:
            emit("WARN", "ROBOT_DELETE_ERROR", robot=robot_name, error=str(e))
            return False

    # ── Assignation groupes / campagnes / médias à un robot ───────────────

    def assign_group_to_robot(self, robot_name: str, group_url: str) -> bool:
        robot_id = self._resolve_robot_id(robot_name)
        if not robot_id:
            return False
        with self._lock:
            conn = self._connect()
            group_id = self._resolve_group_id_conn(conn, group_url)
            conn.execute(
                "INSERT OR IGNORE INTO robot_groups (robot_id, group_id) VALUES (?,?)",
                (robot_id, group_id)
            )
            conn.execute(
                "UPDATE robot_groups SET active=1 WHERE robot_id=? AND group_id=?",
                (robot_id, group_id)
            )
            conn.commit()
            return True

    def assign_campaign_to_robot(self, robot_name: str, campaign_name: str) -> bool:
        robot_id    = self._resolve_robot_id(robot_name)
        campaign_id = self._query_scalar(
            "SELECT id FROM campaigns WHERE name=?", (campaign_name,)
        )
        if not robot_id or not campaign_id:
            return False
        self._exec(
            "INSERT OR IGNORE INTO robot_campaigns (robot_id, campaign_id) VALUES (?,?)",
            (robot_id, campaign_id)
        )
        return True

    def assign_media_to_robot(self, robot_name: str, media_id: int) -> bool:
        robot_id = self._resolve_robot_id(robot_name)
        if not robot_id:
            return False
        self._exec(
            "INSERT OR IGNORE INTO robot_media (robot_id, media_id) VALUES (?,?)",
            (robot_id, media_id)
        )
        return True

    def get_groups_for_robot(self, robot_name: str) -> List[Dict]:
        """Groupes actifs assignés à ce robot."""
        robot_id = self._resolve_robot_id(robot_name)
        if not robot_id:
            return []
        return self._query(
            """SELECT g.* FROM groups g
               JOIN robot_groups rg ON rg.group_id=g.id
               WHERE rg.robot_id=? AND rg.active=1 AND g.active=1
               ORDER BY g.quality_score DESC""",
            (robot_id,)
        )

    def get_campaigns_for_robot(self, robot_name: str) -> List[Dict]:
        robot_id = self._resolve_robot_id(robot_name)
        if not robot_id:
            return []
        return self._query(
            """SELECT c.* FROM campaigns c
               JOIN robot_campaigns rc ON rc.campaign_id=c.id
               WHERE rc.robot_id=? AND c.active=1""",
            (robot_id,)
        )

    def get_media_for_robot(self, robot_name: str) -> List[Dict]:
        robot_id = self._resolve_robot_id(robot_name)
        if not robot_id:
            return []
        return self._query(
            """SELECT m.* FROM media_assets m
               JOIN robot_media rm ON rm.media_id=m.id
               WHERE rm.robot_id=? AND m.active=1""",
            (robot_id,)
        )

    # ══════════════════════════════════════════
    # RUN STATS (anti-doublon runs journaliers)
    # ══════════════════════════════════════════

    def get_run_stats(self, robot_name: str) -> Dict:
        today = datetime.now().date().isoformat()
        row = self._query_one(
            "SELECT * FROM robot_run_stats WHERE robot_name=? AND run_date=?",
            (robot_name, today)
        )
        return dict(row) if row else {
            "robot_name": robot_name, "run_date": today,
            "run_count": 0, "last_run_ts": None,
        }

    def record_run(self, robot_name: str) -> None:
        today = datetime.now().date().isoformat()
        now   = datetime.now().isoformat()
        with self._lock:
            conn = self._connect()
            ex = self._scalar_conn(
                conn,
                "SELECT id FROM robot_run_stats WHERE robot_name=? AND run_date=?",
                (robot_name, today)
            )
            if ex:
                conn.execute(
                    "UPDATE robot_run_stats SET run_count=run_count+1, last_run_ts=?"
                    " WHERE robot_name=? AND run_date=?",
                    (now, robot_name, today)
                )
            else:
                conn.execute(
                    "INSERT INTO robot_run_stats (robot_name,run_date,run_count,last_run_ts)"
                    " VALUES (?,?,1,?)",
                    (robot_name, today, now)
                )
            conn.commit()

    def check_run_limits(self, robot_name: str) -> tuple:
        robot = self.get_robot(robot_name)
        if not robot:
            return False, "Robot introuvable en base"
        max_runs   = robot.get("max_runs_per_day", 2)
        cooldown_s = robot.get("cooldown_between_runs_s", 7200)
        stats      = self.get_run_stats(robot_name)
        last_run_ts = stats.get("last_run_ts")
        if last_run_ts:
            try:
                elapsed = (datetime.now() - datetime.fromisoformat(last_run_ts)).total_seconds()
                if elapsed < cooldown_s:
                    remaining_min = round((cooldown_s - elapsed) / 60, 1)
                    return False, f"Cooldown actif — encore {remaining_min} min"
            except (ValueError, TypeError):
                pass
        if stats.get("run_count", 0) >= max_runs:
            return False, f"Limite journalière atteinte ({stats['run_count']}/{max_runs} runs)"
        return True, "OK"

    # ══════════════════════════════════════════
    # ACCOUNTS
    # ══════════════════════════════════════════

    def create_account(self, name, email=None, profile_url=None):
        now = datetime.now().isoformat()
        with self._lock:
            conn = self._connect()
            cur = conn.execute(
                "INSERT INTO accounts (name,email,profile_url,last_activity_date)"
                " VALUES (?,?,?,?)",
                (name, email, profile_url, now)
            )
            conn.commit()
            return cur.lastrowid

    def ensure_account_exists(self, name, email=None, profile_url=None):
        row = self._query_one("SELECT id FROM accounts WHERE name=?", (name,))
        return row["id"] if row else self.create_account(name, email, profile_url)

    def get_account(self, name):
        return self._query_one("SELECT * FROM accounts WHERE name=?", (name,))

    def get_account_by_id(self, account_id):
        return self._query_one("SELECT * FROM accounts WHERE id=?", (account_id,))

    def get_all_accounts(self):
        return self._query("SELECT * FROM accounts ORDER BY updated_at DESC")

    def get_account_status(self, name):
        row = self._query_one("SELECT status FROM accounts WHERE name=?", (name,))
        return row["status"] if row else None

    def update_account_status(self, account, status, reason=None):
        account_id = self._resolve_account_id(account)
        if account_id is None and isinstance(account, str):
            account_id = self.ensure_account_exists(account)
        if account_id:
            self._set_account_status(account_id, status, reason)

    def _set_account_status(self, account_id, status, reason=None):
        now = datetime.now()
        cooldown = None
        if status == "temporarily_blocked":
            cooldown = (now + timedelta(hours=24)).isoformat()
        elif status == "warning":
            cooldown = (now + timedelta(hours=3)).isoformat()
        self._exec(
            "UPDATE accounts SET status=?,status_reason=?,cooldown_until=?,updated_at=?"
            " WHERE id=?",
            (status, reason, cooldown, now.isoformat(), account_id)
        )

    def can_account_post(self, account, max_per_hour=5):
        row = (self.get_account(account) if isinstance(account, str)
               else self.get_account_by_id(account))
        if not row:
            return False, "Compte inexistant en base"
        cooldown = row.get("cooldown_until")
        if cooldown:
            try:
                until = datetime.fromisoformat(cooldown)
                if datetime.now() < until:
                    remaining_h = int((until - datetime.now()).total_seconds() // 3600)
                    return False, f"Cooldown actif — encore {remaining_h}h"
            except ValueError:
                pass
        status = row.get("status", "healthy")
        if status == "temporarily_blocked":
            return False, "Compte temporairement bloqué"
        if status == "session_expired":
            return False, "Session expirée"
        account_id = row["id"]
        today = datetime.now().date().isoformat()
        posts_today = self._query_scalar(
            "SELECT COUNT(*) FROM publications WHERE account_id=? AND DATE(created_at)=?",
            (account_id, today)
        ) or 0
        if posts_today >= row.get("max_groups_per_day", 20):
            return False, f"Limite quotidienne ({posts_today}/{row.get('max_groups_per_day',20)})"
        one_hour_ago = (datetime.now() - timedelta(hours=1)).isoformat()
        posts_last_hour = self._query_scalar(
            "SELECT COUNT(*) FROM publications"
            " WHERE account_id=? AND created_at>=? AND status='success'",
            (account_id, one_hour_ago)
        ) or 0
        if posts_last_hour >= max_per_hour:
            return False, f"Limite horaire ({posts_last_hour}/{max_per_hour}/h)"
        return True, "OK"

    def record_account_block(self, account, block_type, reason):
        account_id = self._resolve_account_id(account) or self.ensure_account_exists(account)
        self._set_account_status(account_id, "temporarily_blocked", reason)
        self._exec(
            "INSERT INTO account_blocks (account_id,block_type,reason,started_at)"
            " VALUES (?,?,?,?)",
            (account_id, block_type, reason, datetime.now().isoformat())
        )

    def get_account_block_info(self, name):
        row = self.get_account(name)
        if not row or row["status"] != "temporarily_blocked":
            return None
        cooldown = row.get("cooldown_until")
        can_resume = True
        if cooldown:
            try:
                can_resume = datetime.now() >= datetime.fromisoformat(cooldown)
            except Exception:
                pass
        return {
            "status": row["status"], "reason": row.get("status_reason"),
            "until": cooldown, "can_resume": can_resume,
        }

    def mark_warmup_completed(self, account):
        account_id = self._resolve_account_id(account)
        if account_id:
            self._exec(
                "UPDATE accounts SET warmup_completed=1,updated_at=? WHERE id=?",
                (datetime.now().isoformat(), account_id)
            )

    def get_health_score(self, account):
        row = (self.get_account(account) if isinstance(account, str)
               else self.get_account_by_id(account))
        return (row or {}).get("health_score", 100)

    # ══════════════════════════════════════════
    # GROUPS
    # ══════════════════════════════════════════

    def add_group(self, url, name=None, category=None, language="fr", members_count=None):
        with self._lock:
            conn = self._connect()
            rows = self._query_conn(conn, "SELECT id FROM groups WHERE url=?", (url,))
            if rows:
                conn.execute(
                    "UPDATE groups SET name=COALESCE(?,name),"
                    "category=COALESCE(?,category),language=COALESCE(?,language),"
                    "members_count=COALESCE(?,members_count) WHERE url=?",
                    (name, category, language, members_count, url)
                )
                conn.commit()
                return rows[0]["id"]
            cur = conn.execute(
                "INSERT INTO groups (url,name,category,language,members_count)"
                " VALUES (?,?,?,?,?)",
                (url, name, category, language, members_count)
            )
            conn.commit()
            return cur.lastrowid

    def get_group_by_url(self, url):
        return self._query_one("SELECT * FROM groups WHERE url=?", (url,))

    def get_all_groups(self):
        return self._query("SELECT * FROM groups WHERE active=1 ORDER BY quality_score DESC")

    # ══════════════════════════════════════════
    # CAMPAIGNS + VARIANTS
    # ══════════════════════════════════════════

    def upsert_campaign(self, name, description="", language="fr", active=True):
        with self._lock:
            conn = self._connect()
            row = self._query_conn(conn, "SELECT id FROM campaigns WHERE name=?", (name,))
            if row:
                conn.execute(
                    "UPDATE campaigns SET description=?,language=?,active=? WHERE name=?",
                    (description, language, 1 if active else 0, name)
                )
                conn.commit()
                return row[0]["id"]
            cur = conn.execute(
                "INSERT INTO campaigns (name,description,language,active) VALUES (?,?,?,?)",
                (name, description, language, 1 if active else 0)
            )
            conn.commit()
            return cur.lastrowid

    def upsert_variant(self, campaign_id, variant_key, text_fr=None, text_en=None,
                       text_ar=None, cta=None, weight=1, bg_color=None,
                       post_type="text_image"):
        with self._lock:
            conn = self._connect()
            row = self._query_conn(
                conn,
                "SELECT id FROM campaign_variants WHERE campaign_id=? AND variant_key=?",
                (campaign_id, variant_key)
            )
            if row:
                conn.execute(
                    "UPDATE campaign_variants SET text_fr=?,text_en=?,text_ar=?,"
                    "cta=?,weight=?,bg_color=?,post_type=? WHERE id=?",
                    (text_fr, text_en, text_ar, cta, weight, bg_color, post_type, row[0]["id"])
                )
                conn.commit()
                return row[0]["id"]
            cur = conn.execute(
                "INSERT INTO campaign_variants"
                " (campaign_id,variant_key,text_fr,text_en,text_ar,cta,weight,bg_color,post_type)"
                " VALUES (?,?,?,?,?,?,?,?,?)",
                (campaign_id, variant_key, text_fr, text_en, text_ar,
                 cta, weight, bg_color, post_type)
            )
            conn.commit()
            return cur.lastrowid

    def get_campaign_by_name(self, name):
        return self._query_one("SELECT * FROM campaigns WHERE name=? AND active=1", (name,))

    def get_all_campaigns(self):
        return self._query("SELECT * FROM campaigns WHERE active=1 ORDER BY name")

    def get_variants(self, campaign_id):
        return self._query(
            "SELECT * FROM campaign_variants WHERE campaign_id=? ORDER BY weight DESC",
            (campaign_id,)
        )

    def get_seed_group_payloads(self) -> List[Dict]:
        """Retourne les groupes de test par défaut."""
        return [
            {
                "url": "https://www.facebook.com/groups/1616683065262089",
                "name": "Groupe fourni",
                "category": "test",
                "language": "fr",
                "members_count": None,
            },
            {
                "url": "https://www.facebook.com/groups/bon-test-group-1/",
                "name": "BON Test Group 1",
                "category": "test",
                "language": "fr",
                "members_count": 1250,
            },
            {
                "url": "https://www.facebook.com/groups/bon-test-group-2/",
                "name": "BON Test Group 2",
                "category": "test",
                "language": "fr",
                "members_count": 840,
            },
            {
                "url": "https://www.facebook.com/groups/bon-test-group-3/",
                "name": "BON Test Group 3",
                "category": "test",
                "language": "en",
                "members_count": 560,
            },
        ]

    def get_seed_campaign_payloads(self) -> Dict[str, Dict]:
        """Retourne les campagnes de test par défaut."""
        return {
            "test-launch": {
                "name": "test-launch",
                "description": "Campagne de test BON",
                "language": "fr",
                "active": True,
                "variants": [
                    {
                        "id": "v1",
                        "text_fr": "Bonjour, voici un post de test pour valider le flux BON.",
                        "text_en": "Hello, this is a test post to validate the BON flow.",
                        "text_ar": "مرحبا، هذا منشور اختبار للتحقق من تدفق BON.",
                        "cta": "Découvrir",
                        "weight": 2,
                        "bg_color": "#1877F2",
                        "post_type": "text_image",
                    },
                    {
                        "id": "v2",
                        "text_fr": "Version alternative de test pour vérifier les variantes.",
                        "text_en": "Alternative test version to verify variants.",
                        "text_ar": "نسخة اختبار بديلة للتحقق من المتغيرات.",
                        "cta": "Voir plus",
                        "weight": 1,
                        "bg_color": "#42B72A",
                        "post_type": "text",
                    },
                ],
            },
            "follow-up": {
                "name": "follow-up",
                "description": "Campagne de suivi BON",
                "language": "fr",
                "active": True,
                "variants": [
                    {
                        "id": "v1",
                        "text_fr": "Merci pour votre retour, voici un second message de test.",
                        "text_en": "Thanks for your feedback, here is a second test message.",
                        "text_ar": "شكرًا لملاحظاتك، هذه رسالة اختبار ثانية.",
                        "cta": "Répondre",
                        "weight": 1,
                        "bg_color": "#F28C28",
                        "post_type": "text_image",
                    }
                ],
            },
        }

    def bootstrap_default_workspace(self, force: bool = False) -> Dict[str, int]:
        """
        Initialise des données de test SQLite si la base est vide.

        Cette méthode remplace le besoin de ressaisir les groupes ou textes dans
        la console. Elle reste idempotente grâce aux upserts.
        """
        stats = {
            "accounts": 0,
            "robots": 0,
            "groups": 0,
            "campaigns": 0,
            "assignments": 0,
        }

        if str(self.db_path) == ":memory:" and not force:
            return stats

        if not force and str(self.db_path) == ":memory:":
            return stats

        seed_robots = [
            {
                "robot_name": "robot1",
                "account_name": "robot1_account",
                "storage_state_path": str(pathlib.Path("chrome_profiles") / "robot1" / "storage_state.json"),
                "config": {
                    "delay_between_groups": [45, 90],
                    "max_groups_per_run": 5,
                    "max_groups_per_hour": 3,
                    "max_runs_per_day": 1,
                },
                "groups": [
                    "https://www.facebook.com/groups/1616683065262089",
                    "https://www.facebook.com/groups/bon-test-group-1/",
                    "https://www.facebook.com/groups/bon-test-group-2/",
                ],
                "campaigns": ["test-launch"],
            },
            {
                "robot_name": "robot2",
                "account_name": "robot2_account",
                "storage_state_path": str(pathlib.Path("chrome_profiles") / "robot2" / "storage_state.json"),
                "config": {
                    "delay_between_groups": [60, 120],
                    "max_groups_per_run": 4,
                    "max_groups_per_hour": 2,
                    "max_runs_per_day": 1,
                },
                "groups": [
                    "https://www.facebook.com/groups/1616683065262089",
                    "https://www.facebook.com/groups/bon-test-group-2/",
                    "https://www.facebook.com/groups/bon-test-group-3/",
                ],
                "campaigns": ["follow-up"],
            },
        ]

        for group_data in self.get_seed_group_payloads():
            self.add_group(
                url=group_data["url"],
                name=group_data.get("name"),
                category=group_data.get("category"),
                language=group_data.get("language", "fr"),
                members_count=group_data.get("members_count"),
            )
            stats["groups"] += 1

        for camp_key, camp_data in self.get_seed_campaign_payloads().items():
            cid = self.upsert_campaign(
                camp_data.get("name", camp_key),
                camp_data.get("description", ""),
                camp_data.get("language", "fr"),
                camp_data.get("active", True),
            )
            stats["campaigns"] += 1
            for variant in camp_data.get("variants", []):
                self.upsert_variant(
                    campaign_id=cid,
                    variant_key=variant.get("id", "v1"),
                    text_fr=variant.get("text_fr"),
                    text_en=variant.get("text_en"),
                    text_ar=variant.get("text_ar"),
                    cta=variant.get("cta", ""),
                    weight=variant.get("weight", 1),
                    bg_color=variant.get("bg_color"),
                    post_type=variant.get("post_type", "text_image"),
                )

        for robot_data in seed_robots:
            self.ensure_account_exists(robot_data["account_name"])
            self.upsert_robot(
                robot_name=robot_data["robot_name"],
                account_name=robot_data["account_name"],
                storage_state_path=robot_data["storage_state_path"],
                config=robot_data["config"],
            )
            stats["robots"] += 1
            for group_url in robot_data.get("groups", []):
                if self.assign_group_to_robot(robot_data["robot_name"], group_url):
                    stats["assignments"] += 1
            for campaign_name in robot_data.get("campaigns", []):
                if self.assign_campaign_to_robot(robot_data["robot_name"], campaign_name):
                    stats["assignments"] += 1

        emit(
            "INFO",
            "WORKSPACE_BOOTSTRAPPED",
            robots=stats["robots"],
            groups=stats["groups"],
            campaigns=stats["campaigns"],
            assignments=stats["assignments"],
        )
        return stats

    def pick_random_variant(self, campaign_name, language="fr"):
        """Tire un variant pondéré par weight."""
        import random
        camp = self.get_campaign_by_name(campaign_name)
        if not camp:
            return None
        variants = self.get_variants(camp["id"])
        if not variants:
            return None
        weights = [max(1, v.get("weight", 1)) for v in variants]
        chosen  = dict(random.choices(variants, weights=weights, k=1)[0])
        lang_f  = f"text_{language}" if language in ("fr", "en", "ar") else "text_fr"
        chosen["text"] = chosen.get(lang_f) or chosen.get("text_fr") or ""
        return chosen

    # ══════════════════════════════════════════
    # MEDIA ASSETS
    # ══════════════════════════════════════════

    def add_media_asset(self, file_path, file_name=None, campaign_id=None,
                        captcha_text=None, description=None) -> int:
        fname = file_name or pathlib.Path(file_path).name
        cur = self._exec(
            "INSERT INTO media_assets"
            " (file_path,file_name,campaign_id,captcha_text,description) VALUES (?,?,?,?,?)",
            (file_path, fname, campaign_id, captcha_text, description)
        )
        return cur.lastrowid

    def get_media_for_campaign(self, campaign_id) -> List[Dict]:
        return self._query(
            "SELECT * FROM media_assets WHERE campaign_id=? AND active=1",
            (campaign_id,)
        )

    def pick_random_media(self, robot_name=None, campaign_id=None,
                          count=1) -> List[Dict]:
        """
        Tire aléatoirement `count` médias pour ce robot / campagne.
        Si le média a un captcha_text, construit le texte final :
            description_random + " " + captcha_text
        Retourne une liste de dicts avec 'file_path', 'final_text'.
        """
        import random
        if robot_name:
            assets = self.get_media_for_robot(robot_name)
        elif campaign_id:
            assets = self.get_media_for_campaign(campaign_id)
        else:
            assets = self._query("SELECT * FROM media_assets WHERE active=1")

        if not assets:
            return []
        chosen = random.sample(assets, min(count, len(assets)))
        result = []
        for a in chosen:
            a = dict(a)
            # Construire le texte final : description (optionnel) + captcha (optionnel)
            parts = []
            if a.get("description"):
                parts.append(a["description"].strip())
            if a.get("captcha_text"):
                parts.append(a["captcha_text"].strip())
            a["final_caption"] = " ".join(parts) if parts else ""
            result.append(a)
        return result

    # ══════════════════════════════════════════
    # COMMENTS
    # ══════════════════════════════════════════

    def add_comment(self, text, robot_name=None) -> int:
        cur = self._exec(
            "INSERT OR IGNORE INTO comments (text,robot_name) VALUES (?,?)",
            (text, robot_name)
        )
        return cur.lastrowid if cur else 0

    def pick_random_comment(self, robot_name=None) -> Optional[str]:
        """Tire un commentaire aléatoire depuis la DB."""
        if robot_name:
            rows = self._query(
                "SELECT id,text FROM comments"
                " WHERE active=1 AND (robot_name=? OR robot_name IS NULL)"
                " ORDER BY RANDOM() LIMIT 1",
                (robot_name,)
            )
        else:
            rows = self._query(
                "SELECT id,text FROM comments WHERE active=1 ORDER BY RANDOM() LIMIT 1"
            )
        if not rows:
            return None
        self._exec(
            "UPDATE comments SET used_count=used_count+1 WHERE id=?",
            (rows[0]["id"],)
        )
        return rows[0]["text"]

    def get_all_comments(self, robot_name=None) -> List[str]:
        if robot_name:
            return [r["text"] for r in self._query(
                "SELECT text FROM comments"
                " WHERE active=1 AND (robot_name=? OR robot_name IS NULL)",
                (robot_name,)
            )]
        return [r["text"] for r in self._query(
            "SELECT text FROM comments WHERE active=1"
        )]

    # ══════════════════════════════════════════
    # PUBLICATIONS + ANTI-DOUBLON
    # ══════════════════════════════════════════

    def was_published_recently(self, robot_name: str, group_url: str,
                                hours: int = 24,
                                campaign_name: str = None,
                                variant_id: str = None) -> bool:
        """
        Anti-doublon strict :
          - Même robot, même groupe, dans les dernières `hours` heures
          - Si campaign_name + variant_id fournis : vérifie exactement ce variant
          - Sinon : vérifie toute publication réussie dans ce groupe
        """
        acc_row = self._query_one(
            """SELECT a.id FROM accounts a
               JOIN robots r ON r.account_id=a.id
               WHERE r.robot_name=?""",
            (robot_name,)
        )
        grp = self.get_group_by_url(group_url)
        if not acc_row or not grp:
            return False
        since = (datetime.now() - timedelta(hours=hours)).isoformat()
        if campaign_name and variant_id:
            count = self._query_scalar(
                """SELECT COUNT(*) FROM publications
                   WHERE robot_name=? AND group_id=? AND status='success'
                   AND campaign_name=? AND variant_id=? AND created_at>=?""",
                (robot_name, grp["id"], campaign_name, variant_id, since)
            ) or 0
        else:
            count = self._query_scalar(
                """SELECT COUNT(*) FROM publications
                   WHERE robot_name=? AND group_id=? AND status='success'
                   AND created_at>=?""",
                (robot_name, grp["id"], since)
            ) or 0
        return count > 0

    def record_publication(self, robot_name, group_url, account_name=None,
                            status="success", post_content=None,
                            campaign_name=None, variant_id=None,
                            images=None, bg_color=None, post_type="text_image",
                            error_message=None, screenshot_path=None) -> int:
        # Résoudre account_name depuis le robot si non fourni
        if not account_name:
            robot = self.get_robot(robot_name)
            account_name = robot["account_name"] if robot else robot_name

        account_id = self.ensure_account_exists(account_name)
        images_json = json.dumps(images, ensure_ascii=False) if images else None
        now = datetime.now().isoformat()

        with self._lock:
            conn = self._connect()
            group_id = self._resolve_group_id_conn(conn, group_url)
            cur = conn.execute(
                """INSERT INTO publications
                   (robot_name,account_id,group_id,campaign_name,variant_id,
                    text_content,images,bg_color,post_type,
                    status,error_message,screenshot_path,published_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (robot_name, account_id, group_id, campaign_name, variant_id,
                 (post_content or "")[:500], images_json, bg_color, post_type,
                 status, error_message, screenshot_path,
                 now if status == "success" else None)
            )
            pub_id = cur.lastrowid
            if status == "success":
                conn.execute(
                    "UPDATE groups SET successful_posts=successful_posts+1,"
                    "total_posts=total_posts+1,last_post_date=? WHERE id=?",
                    (now, group_id)
                )
                conn.execute(
                    "UPDATE accounts SET successful_posts=successful_posts+1,"
                    "total_posts=total_posts+1,consecutive_failures=0,"
                    "health_score=MIN(100,health_score+2),"
                    "last_post_date=?,last_activity_date=?,updated_at=? WHERE id=?",
                    (now, now, now, account_id)
                )
            else:
                conn.execute(
                    "UPDATE groups SET failed_posts=failed_posts+1,"
                    "total_posts=total_posts+1 WHERE id=?", (group_id,)
                )
                conn.execute(
                    "UPDATE accounts SET failed_posts=failed_posts+1,"
                    "total_posts=total_posts+1,"
                    "consecutive_failures=consecutive_failures+1,"
                    "health_score=MAX(0,health_score-5),"
                    "last_activity_date=?,updated_at=? WHERE id=?",
                    (now, now, account_id)
                )
            conn.commit()
            return pub_id

    def get_publications(self, robot_name=None, group_url=None, limit=50):
        where, params = ["1=1"], []
        if robot_name:
            where.append("robot_name=?")
            params.append(robot_name)
        if group_url:
            grp = self.get_group_by_url(group_url)
            if grp:
                where.append("group_id=?")
                params.append(grp["id"])
        params.append(limit)
        return self._query(
            f"SELECT * FROM publications WHERE {' AND '.join(where)}"
            f" ORDER BY created_at DESC LIMIT ?",
            tuple(params)
        )

    def record_published_comment(self, robot_name, group_url, comment_text,
                                  publication_id=None):
        robot = self.get_robot(robot_name)
        account_id = self.ensure_account_exists(
            robot["account_name"] if robot else robot_name
        )
        with self._lock:
            conn = self._connect()
            gid = self._resolve_group_id_conn(conn, group_url) if group_url else None
            conn.execute(
                "INSERT INTO published_comments"
                " (robot_name,account_id,group_id,publication_id,comment_text)"
                " VALUES (?,?,?,?,?)",
                (robot_name, account_id, gid, publication_id, comment_text)
            )
            conn.commit()

    # ══════════════════════════════════════════
    # ERRORS
    # ══════════════════════════════════════════

    def record_error(self, robot_name=None, account_name=None, group_url=None,
                      error_type=None, error_message=None, step=None,
                      selector_key=None, screenshot_path=None,
                      html_snapshot_path=None):
        account_id = self._resolve_account_id(account_name) if account_name else None
        with self._lock:
            conn = self._connect()
            gid = self._resolve_group_id_conn(conn, group_url) if group_url else None
            conn.execute(
                "INSERT INTO errors"
                " (robot_name,account_id,group_id,error_type,error_message,"
                "  step,selector_key,screenshot_path,html_snapshot_path)"
                " VALUES (?,?,?,?,?,?,?,?,?)",
                (robot_name, account_id, gid, error_type, error_message,
                 step, selector_key, screenshot_path, html_snapshot_path)
            )
            conn.commit()

    def get_recent_errors(self, limit=20):
        return self._query(
            "SELECT * FROM errors ORDER BY created_at DESC LIMIT ?", (limit,)
        )

    # ══════════════════════════════════════════
    # SUBSCRIPTIONS
    # ══════════════════════════════════════════

    def mark_subscribed(self, robot_name, group_url):
        robot = self.get_robot(robot_name)
        if not robot:
            return
        account_id = robot["account_id"]
        with self._lock:
            conn = self._connect()
            gid = self._resolve_group_id_conn(conn, group_url)
            conn.execute(
                "INSERT OR REPLACE INTO subscriptions"
                " (account_id,group_id,status,subscribed_at) VALUES (?,?,'subscribed',?)",
                (account_id, gid, datetime.now().isoformat())
            )
            conn.commit()

    def is_subscribed(self, robot_name, group_url) -> bool:
        robot = self.get_robot(robot_name)
        if not robot:
            return False
        grp = self.get_group_by_url(group_url)
        if not grp:
            return False
        row = self._query_one(
            "SELECT status FROM subscriptions WHERE account_id=? AND group_id=?",
            (robot["account_id"], grp["id"])
        )
        return row is not None and row["status"] == "subscribed"

    # ══════════════════════════════════════════
    # CIRCUIT BREAKER (persistant)
    # ══════════════════════════════════════════

    def get_cb_state(self, robot_name):
        row = self._query_one(
            "SELECT * FROM circuit_breaker_state WHERE robot_name=?", (robot_name,)
        )
        return dict(row) if row else {
            "robot_name": robot_name, "state": "closed",
            "failures": 0, "successes": 0, "opened_at": None,
            "recovery_timeout_s": 900, "failure_threshold": 3, "half_open_max_ok": 2,
        }

    def save_cb_state(self, robot_name, state, failures, successes, opened_at,
                      recovery_timeout_s=900, failure_threshold=3, half_open_max_ok=2):
        self._exec(
            """INSERT OR REPLACE INTO circuit_breaker_state
               (robot_name,state,failures,successes,opened_at,
                recovery_timeout_s,failure_threshold,half_open_max_ok,updated_at)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (robot_name, state, failures, successes, opened_at,
             recovery_timeout_s, failure_threshold, half_open_max_ok,
             datetime.now().isoformat())
        )

    # ══════════════════════════════════════════
    # DM QUEUE
    # ══════════════════════════════════════════

    def enqueue_dm(self, robot_name, target_type, target_id,
                    text_content, media_paths=None, scheduled_at=None) -> int:
        cur = self._exec(
            "INSERT INTO dm_queue"
            " (robot_name,target_type,target_id,text_content,media_paths,scheduled_at)"
            " VALUES (?,?,?,?,?,?)",
            (robot_name, target_type, target_id, text_content,
             json.dumps(media_paths, ensure_ascii=False) if media_paths else None,
             scheduled_at)
        )
        return cur.lastrowid

    def get_pending_dms(self, robot_name, limit=20) -> List[Dict]:
        return self._query(
            "SELECT * FROM dm_queue"
            " WHERE robot_name=? AND status='pending'"
            " AND (scheduled_at IS NULL OR scheduled_at<=?)"
            " ORDER BY created_at ASC LIMIT ?",
            (robot_name, datetime.now().isoformat(), limit)
        )

    def update_dm_status(self, dm_id, status, error_msg=None):
        sent = datetime.now().isoformat() if status == "sent" else None
        self._exec(
            "UPDATE dm_queue SET status=?,attempts=attempts+1,sent_at=?,error_msg=?"
            " WHERE id=?",
            (status, sent, error_msg, dm_id)
        )

    # ══════════════════════════════════════════
    # SELECTOR STATS
    # ══════════════════════════════════════════

    def record_selector_attempt(self, selector_key, success,
                                 used_selector=None, failure_reason=None):
        now = datetime.now().isoformat()
        ex = self._query_one(
            "SELECT id,total_attempts,successful_attempts FROM selector_stats"
            " WHERE selector_key=?",
            (selector_key,)
        )
        if ex:
            if success:
                self._exec(
                    "UPDATE selector_stats SET"
                    " successful_attempts=successful_attempts+1,"
                    " total_attempts=total_attempts+1,"
                    " last_success_date=?,"
                    " working_selector=COALESCE(?,working_selector),"
                    " success_rate=(successful_attempts+1)*100.0/(total_attempts+1),"
                    " updated_at=? WHERE selector_key=?",
                    (now, used_selector, now, selector_key)
                )
            else:
                self._exec(
                    "UPDATE selector_stats SET"
                    " failed_attempts=failed_attempts+1,"
                    " total_attempts=total_attempts+1,"
                    " last_failure_date=?,"
                    " last_failure_reason=COALESCE(?,last_failure_reason),"
                    " success_rate=successful_attempts*100.0/(total_attempts+1),"
                    " updated_at=? WHERE selector_key=?",
                    (now, failure_reason, now, selector_key)
                )
        else:
            self._exec(
                "INSERT INTO selector_stats"
                " (selector_key,working_selector,total_attempts,successful_attempts,"
                "  failed_attempts,success_rate,last_success_date,last_failure_date,"
                "  last_failure_reason) VALUES (?,?,1,?,?,?,?,?,?)",
                (selector_key, used_selector if success else None,
                 1 if success else 0, 0 if success else 1,
                 100.0 if success else 0.0,
                 now if success else None,
                 now if not success else None,
                 failure_reason if not success else None)
            )

    def get_selector_stats(self):
        return self._query("SELECT * FROM selector_stats ORDER BY success_rate ASC")

    # ══════════════════════════════════════════
    # CONFIG KV
    # ══════════════════════════════════════════

    def config_set(self, key, value):
        now = datetime.now().isoformat()
        self._exec(
            "INSERT OR REPLACE INTO config_kv (key,value,updated_at) VALUES (?,?,?)",
            (key, value, now)
        )

    def config_get(self, key, default=None):
        row = self._query_one("SELECT value FROM config_kv WHERE key=?", (key,))
        return row["value"] if row else default

    def config_all(self):
        return {r["key"]: r["value"] for r in self._query("SELECT key,value FROM config_kv")}

    # ══════════════════════════════════════════
    # CAPTCHA LOG (v14)
    # ══════════════════════════════════════════

    def log_captcha_event(self, robot_name, solve_type, status, error_message=None) -> None:
        self._exec(
            "INSERT INTO captcha_solve_log (robot_name,solve_type,status,error_message)"
            " VALUES (?,?,?,?)",
            (robot_name, solve_type, status, error_message)
        )

    def get_captcha_solve_stats(self, days: int = 7) -> List[Dict]:
        since = (datetime.now() - timedelta(days=days)).isoformat()
        return self._query(
            "SELECT status, COUNT(*) AS cnt FROM captcha_solve_log"
            " WHERE created_at>=? GROUP BY status",
            (since,)
        )

    # ══════════════════════════════════════════
    # SCHEDULER JOBS (v14)
    # ══════════════════════════════════════════

    def scheduler_upsert_job(
        self, job_id: str, robot_name: str, cron_expression: str,
        command_name: str = "post", active: int = 1,
    ) -> None:
        now = datetime.now().isoformat()
        row = self._query_one("SELECT id FROM scheduler_jobs WHERE job_id=?", (job_id,))
        if row:
            self._exec(
                """UPDATE scheduler_jobs SET robot_name=?, cron_expression=?,
                   command_name=?, active=? WHERE job_id=?""",
                (robot_name, cron_expression, command_name, active, job_id)
            )
        else:
            self._exec(
                """INSERT INTO scheduler_jobs
                   (job_id,robot_name,cron_expression,command_name,active,created_at)
                   VALUES (?,?,?,?,?,?)""",
                (job_id, robot_name, cron_expression, command_name, active, now)
            )

    def scheduler_list_jobs(self) -> List[Dict]:
        return self._query(
            "SELECT * FROM scheduler_jobs ORDER BY robot_name, job_id"
        )

    def scheduler_delete_job(self, job_id: str) -> bool:
        with self._lock:
            conn = self._connect()
            cur = conn.execute("DELETE FROM scheduler_jobs WHERE job_id=?", (job_id,))
            conn.commit()
            return cur.rowcount > 0

    def scheduler_set_active(self, job_id: str, active: int) -> bool:
        with self._lock:
            conn = self._connect()
            cur = conn.execute(
                "UPDATE scheduler_jobs SET active=? WHERE job_id=?",
                (active, job_id)
            )
            conn.commit()
            return cur.rowcount > 0

    def scheduler_update_run_meta(
        self, job_id: str, last_run_at: str, next_run_at: Optional[str] = None
    ) -> None:
        if next_run_at:
            self._exec(
                "UPDATE scheduler_jobs SET last_run_at=?, next_run_at=? WHERE job_id=?",
                (last_run_at, next_run_at, job_id)
            )
        else:
            self._exec(
                "UPDATE scheduler_jobs SET last_run_at=? WHERE job_id=?",
                (last_run_at, job_id)
            )

    # ══════════════════════════════════════════
    # EXPORT / PAGINATION PUBLICATIONS (v14)
    # ══════════════════════════════════════════

    def _publication_export_rows(
        self,
        robot_name: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> List[Dict]:
        """Lignes brutes pour les exports CSV/XLSX avec filtres optionnels."""
        sql_base = """SELECT p.id, p.robot_name, a.name AS account, g.url AS group_url,
                        g.name AS group_name, p.campaign_name, p.variant_id, p.post_type,
                        p.status, p.text_content, p.created_at, p.published_at,
                        p.error_message
                 FROM publications p
                 JOIN accounts a ON a.id = p.account_id
                 JOIN groups g ON g.id = p.group_id"""
        conditions = []
        params: list = []
        if robot_name:
            conditions.append("p.robot_name = ?")
            params.append(robot_name)
        if date_from:
            conditions.append("p.created_at >= ?")
            params.append(date_from)
        if date_to:
            date_to_end = date_to if "T" in date_to else date_to + "T23:59:59"
            conditions.append("p.created_at <= ?")
            params.append(date_to_end)
        where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
        return self._query(sql_base + where + " ORDER BY p.created_at DESC", tuple(params))

    def export_publications_csv(
        self,
        out_path,
        robot_name: Optional[str] = None,
        encoding: str = "utf-8",
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> int:
        rows = self._publication_export_rows(robot_name, date_from=date_from, date_to=date_to)
        path = pathlib.Path(out_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        # CSV historique : sous-ensemble stable pour Excel / outils tiers
        fieldnames = [
            "id", "robot_name", "account", "group_url",
            "campaign_name", "variant_id", "status", "created_at", "error_message",
        ]
        with open(path, "w", newline="", encoding=encoding) as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow(dict(r))
        return len(rows)

    def export_publications_xlsx(
        self,
        out_path,
        robot_name: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> int:
        """Export enrichi (colonnes completes). Necessite : pip install openpyxl"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise ImportError("Installez openpyxl : pip install openpyxl") from e

        rows = self._publication_export_rows(robot_name, date_from=date_from, date_to=date_to)
        headers = [
            "id", "robot_name", "account", "group_url", "group_name",
            "campaign_name", "variant_id", "post_type", "status",
            "text_content", "created_at", "published_at", "error_message",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "publications"
        hdr_fill = PatternFill("solid", fgColor="1877F2")
        hdr_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for ri, row in enumerate(rows, 2):
            d = dict(row)
            for ci, h in enumerate(headers, 1):
                v = d.get(h)
                if v is not None and not isinstance(v, (int, float, bool)):
                    v = str(v)[:32000]
                ws.cell(row=ri, column=ci, value=v)
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = min(28, max(12, len(headers[i - 1]) + 2))
        ws.freeze_panes = "A2"
        if hasattr(out_path, "write"):
            wb.save(out_path)
        else:
            path = pathlib.Path(out_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(path))
        return len(rows)

    def get_publications_paginated(
        self,
        limit: int = 50,
        offset: int = 0,
        robot_name: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> List[Dict]:
        """Retourne les publications paginées avec filtres optionnels.

        Args:
            limit: Nombre max de résultats (max 500).
            offset: Décalage pour la pagination.
            robot_name: Filtre par robot (None = tous les robots).
            date_from: Filtre date début ISO 8601, ex. '2026-01-01' (inclus).
            date_to: Filtre date fin ISO 8601, ex. '2026-03-31' (inclus jusqu'à 23:59:59).
        """
        conditions = []
        params: list = []
        if robot_name:
            conditions.append("p.robot_name = ?")
            params.append(robot_name)
        if date_from:
            conditions.append("p.created_at >= ?")
            params.append(date_from)
        if date_to:
            # Inclut toute la journée de date_to
            date_to_end = date_to if "T" in date_to else date_to + "T23:59:59"
            conditions.append("p.created_at <= ?")
            params.append(date_to_end)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        params.extend([limit, offset])
        return self._query(
            f"""SELECT p.*, g.url AS group_url FROM publications p
               JOIN groups g ON g.id = p.group_id
               {where}
               ORDER BY p.created_at DESC LIMIT ? OFFSET ?""",
            tuple(params)
        )

    # ══════════════════════════════════════════
    # DASHBOARD
    # ══════════════════════════════════════════

    def get_dashboard_stats(self) -> Dict:
        today = datetime.now().date().isoformat()
        return {
            "total_robots":          self._query_scalar("SELECT COUNT(*) FROM robots WHERE active=1") or 0,
            "total_accounts":        self._query_scalar("SELECT COUNT(*) FROM accounts") or 0,
            "healthy_accounts":      self._query_scalar("SELECT COUNT(*) FROM accounts WHERE status='healthy'") or 0,
            "blocked_accounts":      self._query_scalar("SELECT COUNT(*) FROM accounts WHERE status='temporarily_blocked'") or 0,
            "total_groups":          self._query_scalar("SELECT COUNT(*) FROM groups WHERE active=1") or 0,
            "total_campaigns":       self._query_scalar("SELECT COUNT(*) FROM campaigns WHERE active=1") or 0,
            "total_media_assets":    self._query_scalar("SELECT COUNT(*) FROM media_assets WHERE active=1") or 0,
            "posts_today":           self._query_scalar("SELECT COUNT(*) FROM publications WHERE DATE(created_at)=?", (today,)) or 0,
            "successful_posts_today":self._query_scalar("SELECT COUNT(*) FROM publications WHERE DATE(created_at)=? AND status='success'", (today,)) or 0,
            "failed_posts_today":    self._query_scalar("SELECT COUNT(*) FROM publications WHERE DATE(created_at)=? AND status='failed'", (today,)) or 0,
            "pending_dms":           self._query_scalar("SELECT COUNT(*) FROM dm_queue WHERE status='pending'") or 0,
            "errors_today":          self._query_scalar("SELECT COUNT(*) FROM errors WHERE DATE(created_at)=DATE('now')") or 0,
            "total_comments_bank":   self._query_scalar("SELECT COUNT(*) FROM comments WHERE active=1") or 0,
        }

    # ══════════════════════════════════════════
    # IMPORT JSON → SQL (migration one-shot)
    # ══════════════════════════════════════════

    def import_campaigns_from_json(self, path) -> int:
        p = pathlib.Path(path)
        if not p.exists():
            return 0
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            emit("WARN", "CAMPAIGN_JSON_IMPORT_ERROR", error=str(e))
            return 0
        count = 0
        for camp_key, camp_data in data.get("campaigns", {}).items():
            name = camp_data.get("name", camp_key)
            cid  = self.upsert_campaign(name, camp_data.get("description", ""),
                                        camp_data.get("language", "fr"),
                                        camp_data.get("active", True))
            for v in camp_data.get("variants", []):
                self.upsert_variant(
                    campaign_id = cid,
                    variant_key = v.get("id", "v1"),
                    text_fr     = v.get("text_fr"),
                    text_en     = v.get("text_en"),
                    text_ar     = v.get("text_ar"),
                    cta         = v.get("cta", ""),
                    weight      = v.get("weight", 1),
                    bg_color    = v.get("bg_color"),
                    post_type   = v.get("post_type", "text_image"),
                )
                for img in v.get("images", []):
                    self.add_media_asset(img, campaign_id=cid)
            count += 1
        emit("INFO", "CAMPAIGNS_IMPORTED", count=count)
        return count

    def import_groups_from_json(self, path) -> int:
        p = pathlib.Path(path)
        if not p.exists():
            return 0
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            emit("WARN", "GROUPS_JSON_IMPORT_ERROR", error=str(e))
            return 0
        count = 0
        for g in data.get("groups", []):
            if g.get("url"):
                self.add_group(url=g["url"], name=g.get("name"),
                               category=g.get("category"),
                               language=g.get("language", "fr"))
                count += 1
        emit("INFO", "GROUPS_IMPORTED", count=count)
        return count


# ── Singleton thread-safe ─────────────────────────────────────────────────────

_db_instance: Optional[BONDatabase] = None
_db_lock = threading.Lock()


def get_database() -> BONDatabase:
    """Retourne l'instance singleton de BONDatabase (creation lazy, thread-safe)."""
    global _db_instance
    if _db_instance is None:
        with _db_lock:
            if _db_instance is None:
                _db_instance = BONDatabase()
    return _db_instance


def reset_database(new_instance: Optional[BONDatabase] = None) -> None:
    """Remplace le singleton — USAGE TESTS UNIQUEMENT.

    Ferme proprement l'ancienne connexion avant le remplacement.
    Appelable avec None pour forcer la recreation au prochain get_database().

    Exemple dans les tests :
        db = BONDatabase(":memory:")
        reset_database(db)
        yield db
        reset_database()     # remet a None
        db.close()           # libere le descripteur
    """
    global _db_instance
    with _db_lock:
        if _db_instance is not None:
            try:
                _db_instance.close()
            except Exception:
                pass
        _db_instance = new_instance
